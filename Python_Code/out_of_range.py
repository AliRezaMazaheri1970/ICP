import sys
import sqlite3
import pandas as pd
import re
import logging
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar, QMessageBox,
    QFileDialog, QLabel, QDialog, QComboBox, QPushButton, QListWidget, QListWidgetItem, QLineEdit, QCheckBox, QGridLayout, QFrame,QAbstractItemView,QButtonGroup
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer
from PyQt6.QtGui import QFont, QPixmap, QColor, QPalette
from pyqtgraph import PlotWidget, mkPen
from PyQt6.QtGui import QFont, QColor
from PyQt6.QtCore import Qt
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import numpy as np
from pathlib import Path
from PIL import Image
import csv
import shutil
import os
from db.db import get_db_connection,get_ver_db,resource_path
from utils.date import extract_date,validate_jalali_date
from utils.utils import normalize_crm_id
from utils.var_main import BLANK_PATTERN
from db.qc_queries import get_verification_value,load_crm_data_for_file
# Setup logging with UTF-8 encoding
log_file = Path("crm_visualizer.log").resolve()
file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.DEBUG)
console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
logger.handlers = []
logger.addHandler(file_handler)
logger.addHandler(console_handler)

class OutOfRangeThread(QThread):
    out_of_range_data = pyqtSignal(pd.DataFrame)
    progress_updated = pyqtSignal(int)

    def __init__(self, db_path, file_name, percentage, ver_db_path):
        super().__init__()
        self.db_path = db_path
        self.file_name = file_name
        self.percentage = percentage
        self.ver_db_path = ver_db_path
        self.verification_cache = {} # Initialize verification cache

    def run(self):
        try:
            self.progress_updated.emit(20)
            logger.info(f"Starting OutOfRangeThread for file: {self.file_name}")
            df = load_crm_data_for_file(self.file_name)
            
            crm_df = df[df['crm_id'] != 'BLANK'].copy()
            blank_df = df[df['crm_id'] == 'BLANK'].copy()
            crm_df['norm_crm_id'] = crm_df['crm_id'].apply(normalize_crm_id)
            out_df = pd.DataFrame()
            if crm_df.empty:
                logger.warning(f"No CRM data found for file {self.file_name}")
                self.out_of_range_data.emit(out_df)
                self.progress_updated.emit(100)
                return
            # Extract base elements (e.g., 'Ce' from 'Ce 140')
            crm_df['base_element'] = crm_df['element'].apply(lambda x: x.split()[0] if isinstance(x, str) and ' ' in x else x)
            unique_crms = crm_df['norm_crm_id'].unique()
            unique_base_elements = crm_df['base_element'].unique()
            self.progress_updated.emit(40)
            for crm_id in unique_crms:
                for base_element in unique_base_elements:
                    # Get all wavelengths for this base element and CRM ID
                    element_df = crm_df[(crm_df['norm_crm_id'] == crm_id) & (crm_df['base_element'] == base_element)]
                    if element_df.empty:
                        continue
                    ver_value = self.get_verification_value(crm_id, base_element)
                    if ver_value is None:
                        continue
                    lcl = ver_value * (1 - self.percentage / 100)
                    ucl = ver_value * (1 + self.percentage / 100)
                    # Calculate best wavelength based on corrected_value or value
                    best_row = None
                    min_diff = float('inf')
                    for _, row in element_df.iterrows():
                        value = row['value']
                        blank_value, corrected_value = self.select_best_blank(row, blank_df, ver_value)
                        # Use corrected_value if available, else use value
                        target_value = corrected_value if pd.notna(corrected_value) else value
                        diff = abs(target_value - ver_value)
                        if diff < min_diff:
                            min_diff = diff
                            best_row = {
                                'crm_id': row['crm_id'],
                                'element': row['element'],
                                'value': value,
                                'corrected_value': corrected_value if pd.notna(corrected_value) else pd.NA,
                                'ref_value': ver_value,
                                'out_no_blank': not (lcl <= value <= ucl),
                                'out_with_blank': not (lcl <= corrected_value <= ucl) if pd.notna(corrected_value) else not (lcl <= value <= ucl)
                            }
                    if best_row and (best_row['out_no_blank'] or best_row['out_with_blank']):
                        out_df = pd.concat([out_df, pd.DataFrame([best_row])], ignore_index=True)
                        logger.info(f"Added out-of-range record for CRM {crm_id}, Element {best_row['element']}: {best_row}")
            self.progress_updated.emit(100)
            self.out_of_range_data.emit(out_df)
        except Exception as e:
            logger.error(f"Error computing out of range for {self.file_name}: {str(e)}")
            self.out_of_range_data.emit(pd.DataFrame())
            self.progress_updated.emit(100)

    def get_verification_value(self, crm_id, element):
        return get_verification_value(crm_id, element)
    
    def select_best_blank(self, crm_row, blank_df, ver_value):
        if blank_df.empty or ver_value is None:
            logger.debug(f"No blank correction applied: empty blank_df={blank_df.empty}, ver_value={ver_value}")
            return None, crm_row['value']
        relevant_blanks = blank_df[
            (blank_df['file_name'] == crm_row['file_name']) &
            (blank_df['folder_name'] == crm_row['folder_name']) &
            (blank_df['element'] == crm_row['element'])
        ]
        if relevant_blanks.empty:
            logger.debug(f"No relevant blanks found for CRM: file={crm_row['file_name']}, folder={crm_row['folder_name']}, element={crm_row['element']}")
            return None, crm_row['value']
        blank_valid_pattern = re.compile(BLANK_PATTERN, re.IGNORECASE)
        valid_blanks = relevant_blanks[relevant_blanks['solution_label'].apply(lambda x: bool(blank_valid_pattern.match(str(x).strip())))]
        if valid_blanks.empty:
            logger.debug(f"No valid blanks found for CRM row {crm_row['id']}")
            return None, crm_row['value']
        initial_diff = abs(crm_row['value'] - ver_value)
        best_blank_value = None
        best_diff = initial_diff
        corrected_value = crm_row['value']
        for _, blank_row in valid_blanks.iterrows():
            blank_value = blank_row['value']
            if pd.notna(blank_value):
                try:
                    corrected = crm_row['value'] - blank_value
                    new_diff = abs(ver_value - corrected)
                    logger.debug(f"Blank: solution_label={blank_row['solution_label']}, value={blank_value:.2f}, corrected={corrected:.2f}, new_diff={new_diff:.2f}, initial_diff={initial_diff:.2f}")
                    if new_diff < initial_diff:
                        best_diff = new_diff
                        best_blank_value = blank_value
                        corrected_value = corrected
                except (TypeError, ValueError) as e:
                    logger.warning(f"Invalid blank value {blank_value} for CRM row {crm_row['id']}: {str(e)}")
                    continue
        if best_blank_value is not None:
            logger.info(f"Selected blank value {best_blank_value:.2f} for CRM row {crm_row['id']}, corrected value={corrected_value:.2f}, diff={best_diff:.2f}")
        else:
            logger.debug(f"No valid blank value selected for CRM row {crm_row['id']}, using original value={crm_row['value']:.2f}")
        return best_blank_value, corrected_value


class OutOfRangeFilesDialog(QDialog):
    def __init__(self, parent=None, file_names=[], db_path=None, percentage=10.0, ver_db_path=None):
        super().__init__(parent)
        self.setWindowTitle("Out of Range Elements")
        self.setFixedSize(400, 400)
        self.db_path = db_path
        self.ver_db_path = ver_db_path
        self.percentage = percentage
      
        self.layout = QVBoxLayout()
        self.label = QLabel("Select a file to view out-of-range elements:")
        self.file_list = QListWidget()
      
        for file_name in sorted(set(file_names)):
            item = QListWidgetItem(file_name)
            item.setData(32, file_name)
            self.file_list.addItem(item)
      
        self.layout.addWidget(self.label)
        self.layout.addWidget(self.file_list)
        self.setLayout(self.layout)
      
        self.file_list.itemClicked.connect(self.on_file_clicked)

    def on_file_clicked(self, item):
        file_name = item.data(32)
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setVisible(True)
        self.layout.addWidget(self.progress_bar)
      
        self.out_of_range_thread = OutOfRangeThread(self.db_path, file_name, self.percentage, self.ver_db_path)
        self.out_of_range_thread.out_of_range_data.connect(self.on_out_of_range_data)
        self.out_of_range_thread.progress_updated.connect(self.progress_bar.setValue)
        self.out_of_range_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
        self.out_of_range_thread.start()

    def on_out_of_range_data(self, out_df):
        dialog = OutOfRangeTableDialog(self, out_df)
        dialog.exec()

class OutOfRangeTableDialog(QDialog):
    def __init__(self, parent=None, out_df=None):
        super().__init__(parent)
        self.setWindowTitle("Out of Range Elements")
        self.setMinimumSize(800, 500)
        self.out_df = out_df
        self.layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(6) # Added column for Percentage Difference
        self.table_widget.setHorizontalHeaderLabels([
            "CRM ID", "Element", "Value", "Corrected Value", "Ref Value", "Diff %"
        ])
        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        header.setMinimumSectionSize(100)
        header.setDefaultSectionSize(150)
        header.setStretchLastSection(True)
        header.setFont(QFont("Segoe UI", 12, QFont.Weight.Bold))
        vertical_header = self.table_widget.verticalHeader()
        vertical_header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        vertical_header.setDefaultSectionSize(40)
        vertical_header.setFont(QFont("Segoe UI", 10))
        if out_df is not None and not out_df.empty:
            self.table_widget.setRowCount(len(out_df))
            for i, row in out_df.iterrows():
                crm_id_item = QTableWidgetItem(str(row['crm_id']) if pd.notna(row['crm_id']) else "")
                element_item = QTableWidgetItem(str(row['element']) if pd.notna(row['element']) else "")
                value_item = QTableWidgetItem(f"{row['value']:.6f}" if pd.notna(row['value']) else "")
                corrected_item = QTableWidgetItem(f"{row['corrected_value']:.6f}" if pd.notna(row['corrected_value']) else "")
                ref_item = QTableWidgetItem(f"{row['ref_value']:.6f}" if pd.notna(row['ref_value']) else "")
              
                # Calculate percentage difference
                percentage_diff = (abs(row['corrected_value'] - row['ref_value']) / row['ref_value'] * 100
                                 if pd.notna(row['corrected_value']) and pd.notna(row['ref_value']) and row['ref_value'] != 0
                                 else 0)
                diff_item = QTableWidgetItem(f"{percentage_diff:.2f}%")
                value_color = QColor('red') if row['out_no_blank'] else QColor('green')
                corrected_color = QColor('red') if row['out_with_blank'] else QColor('green')
                value_item.setForeground(value_color)
                corrected_item.setForeground(corrected_color)
                diff_item.setForeground(QColor('black'))
                self.table_widget.setItem(i, 0, crm_id_item)
                self.table_widget.setItem(i, 1, element_item)
                self.table_widget.setItem(i, 2, value_item)
                self.table_widget.setItem(i, 3, corrected_item)
                self.table_widget.setItem(i, 4, ref_item)
                self.table_widget.setItem(i, 5, diff_item)
        else:
            self.table_widget.setRowCount(0)
        self.table_widget.resizeColumnsToContents()
        self.table_widget.update()
        self.table_widget.repaint()
        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        self.layout.addWidget(self.table_widget)
        self.layout.addWidget(self.export_button)
        self.setLayout(self.layout)

    def export_to_excel(self):
        if self.out_df is None or self.out_df.empty:
            QMessageBox.warning(self, "Warning", "No data to export")
            return
        fname, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if fname:
            try:
                # Prepare DataFrame for export
                export_df = self.out_df[['crm_id', 'element', 'value', 'corrected_value', 'ref_value']].copy()
                # Add percentage difference column (store as decimal, e.g., 0.5530 for 55.30%)
                export_df['Diff %'] = export_df.apply(
                    lambda row: (abs(row['corrected_value'] - row['ref_value']) / row['ref_value']
                                 if pd.notna(row['corrected_value']) and pd.notna(row['ref_value']) and row['ref_value'] != 0
                                 else 0),
                    axis=1
                )
                # Save to Excel with openpyxl
                with pd.ExcelWriter(fname, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='OutOfRange')
                    workbook = writer.book
                    worksheet = writer.sheets['OutOfRange']
                    # Define styles
                    red_font = Font(color='FF0000', bold=True)
                    green_font = Font(color='008000', bold=True)
                    black_font = Font(color='000000')
                    header_font = Font(bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
                    out_of_range_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    center_align = Alignment(horizontal='center', vertical='center')
                    # Apply header styles
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border
                    # Apply formatting to data rows
                    for row_idx, row in enumerate(self.out_df.itertuples(), start=2):
                        value_out = row.out_no_blank
                        corrected_out = row.out_with_blank
                        # Apply styles to cells
                        for col_idx, col_name in enumerate(['crm_id', 'element', 'value', 'corrected_value', 'ref_value', 'Diff %'], 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                            cell.alignment = center_align
                          
                            if col_name == 'value':
                                cell.font = red_font if value_out else green_font
                                cell.number_format = '0.000000'
                            elif col_name == 'corrected_value':
                                cell.font = red_font if corrected_out else green_font
                                cell.number_format = '0.000000'
                            elif col_name == 'ref_value':
                                cell.font = black_font
                                cell.number_format = '0.000000'
                            elif col_name == 'Diff %':
                                cell.font = black_font
                                cell.number_format = '0.00%' # Format as percentage (0.5530 → 55.30%)
                            else:
                                cell.font = black_font
                            # Apply row background if out of range
                            if value_out or corrected_out:
                                cell.fill = out_of_range_fill
                    # Adjust column widths
                    for col in worksheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30) # Cap width at 30 for aesthetics
                        worksheet.column_dimensions[column].width = adjusted_width
                    # Add filters to the table
                    worksheet.auto_filter.ref = worksheet.dimensions
                QMessageBox.information(self, "Success", f"Data exported to {fname}")
                logger.info(f"Exported out-of-range data to {fname}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")
                logger.error(f"Error exporting Excel: {str(e)}")
