from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel, QTableView, QAbstractItemView,
    QHeaderView, QScrollBar, QComboBox, QLineEdit, QDialog, QFileDialog, QMessageBox, QGroupBox, QProgressBar, QProgressDialog,
    QTabWidget, QScrollArea, QCheckBox
)
from PyQt6.QtCore import Qt, QAbstractTableModel, QTimer, QThread, pyqtSignal, pyqtSlot
from PyQt6.QtGui import QStandardItemModel, QStandardItem, QFont, QColor
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import numpy as np
import os
import platform
import math
from functools import reduce
import re
import logging

from .changeReport import ChangesReportDialog
from ..Common.column_filter import ColumnFilterDialog
from ..Common.Freeze_column import FreezeTableWidget

# Setup logging
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Global stylesheet
global_style = """
    QWidget {
        background-color: #F5F7FA;
        font-family: 'Inter', 'Segoe UI', sans-serif;
        font-size: 13px;
    }
    QGroupBox {
        font-weight: bold;
        color: #1A3C34;
        margin-top: 15px;
        border: 1px solid #D0D7DE;
        border-radius: 6px;
        padding: 10px;
    }
    QGroupBox::title {
        subcontrol-origin: margin;
        subcontrol-position: top left;
        padding: 0 5px;
        left: 10px;
    }
    QPushButton {
        background-color: #2E7D32;
        color: white;
        border: none;
        padding: 8px 16px;
        font-weight: 600;
        font-size: 13px;
        border-radius: 6px;
    }
    QPushButton:hover {
        background-color: #1B5E20;
    }
    QPushButton:disabled {
        background-color: #E0E0E0;
        color: #6B7280;
    }
    QTableView {
        background-color: #FFFFFF;
        border: 1px solid #D0D7DE;
        gridline-color: #E5E7EB;
        font-size: 12px;
        selection-background-color: #DBEAFE;
        selection-color: #1A3C34;
    }
    QHeaderView::section {
        background-color: #F9FAFB;
        font-weight: 600;
        color: #1A3C34;
        border: 1px solid #D0D7DE;
        padding: 6px;
    }
    QTableView::item:selected {
        background-color: #DBEAFE;
        color: #1A3C34;
    }
    QTableView::item {
        padding: 0px;
    }
    QLineEdit {
        padding: 6px;
        border: 1px solid #D0D7DE;
        border-radius: 4px;
        font-size: 13px;
    }
    QLineEdit:focus {
        border: 1px solid #2E7D32;
    }
    QLabel {
        font-size: 13px;
        color: #1A3C34;
    }
    QComboBox {
        padding: 6px;
        border: 1px solid #D0D7DE;
        border-radius: 4px;
        font-size: 13px;
    }
    QComboBox:focus {
        border: 1px solid #2E7D32;
    }
    QProgressBar {
        border: 1px solid #D0D7DE;
        border-radius: 4px;
        text-align: center;
    }
    QProgressBar::chunk {
        background-color: #2E7D32;
    }
"""

class PandasModel(QAbstractTableModel):
    """Custom model to display pandas DataFrame in QTableView"""
    def __init__(self, data=pd.DataFrame(), format_value=None):
        super().__init__()
        self._data = data
        self._format_value = format_value
        self._checkboxes = [False] * len(data)

    def rowCount(self, parent=None):
        return len(self._data)

    def columnCount(self, parent=None):
        return len(self._data.columns) + 1

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        col = index.column()
        row = index.row()
        if col == 0:
            if role == Qt.ItemDataRole.CheckStateRole:
                return Qt.CheckState.Checked if self._checkboxes[row] else Qt.CheckState.Unchecked
            return None
        else:
            value = self._data.iloc[row, col - 1]
            if role == Qt.ItemDataRole.DisplayRole:
                if self._format_value is not None:
                    return self._format_value(value)
                return str(value)
            elif role == Qt.ItemDataRole.BackgroundRole:
                if self._checkboxes[row]:
                    return QColor("#E6F3FA")
                return QColor("#F9FAFB") if row % 2 else Qt.GlobalColor.white
        return None

    def setData(self, index, value, role=Qt.ItemDataRole.EditRole):
        if not index.isValid():
            return False
        row = index.row()
        col = index.column()
        if col == 0 and role == Qt.ItemDataRole.CheckStateRole:
            self._checkboxes[row] = (value == Qt.CheckState.Checked.value)
            self.dataChanged.emit(index, index)
            row_start = self.index(row, 0)
            row_end = self.index(row, self.columnCount() - 1)
            self.dataChanged.emit(row_start, row_end)
            return True
        if role == Qt.ItemDataRole.EditRole and col > 0:
            try:
                numeric_value = float(value)
                df_col_index = col - 1
                if df_col_index >= len(self._data.columns):
                    return False
                self._data.iloc[row, df_col_index] = numeric_value
                self.dataChanged.emit(index, index)
                return True
            except (ValueError, IndexError):
                return False
        return False

    def flags(self, index):
        flags = super().flags(index)
        if index.column() == 0:
            flags |= Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled
        return flags

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole:
            if orientation == Qt.Orientation.Horizontal:
                if section == 0:
                    return "Select"
                return str(self._data.columns[section - 1]) if section - 1 < len(self._data.columns) else ""
            return str(self._data.index[section]) if section < len(self._data.index) else ""
        return None

class DataWorker(QThread):
    data_ready = pyqtSignal(pd.DataFrame)
    error_occurred = pyqtSignal(str)

    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent

    def run(self):
        try:
            df = self.parent.compute_filtered_data()
            if df is None:
                self.data_ready.emit(pd.DataFrame())
            else:
                self.data_ready.emit(df)
        except Exception as e:
            logger.error(f"Error in compute_filtered_data: {str(e)}")
            self.error_occurred.emit(str(e))

class ResultsFrame(QWidget):
    def __init__(self, app, parent=None):
        super().__init__(parent)
        self.app = app
        self.setStyleSheet(global_style)
        self.search_var = ""
        self.filter_field = "Solution Label"
        self.filter_values = {}
        self.column_filters = {}
        self.results_df = None
        self.column_widths = {}
        self.column_backups = {}
        self.last_filtered_data = None
        self.last_pivot_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.decimal_places = "1"
        self.data_hash = None
        self.worker = None
        self.instance_id = id(self)
        logger.debug(f"ResultsFrame initialized with instance_id: {self.instance_id}")
        self.setup_ui()
        self.app.notify_data_changed = self.on_data_changed

        # تشخیص نوع داده
        df = self.app.get_data()
        if df is not None and not df.empty:
            if 'Element' in df.columns and 'Corr Con' in df.columns:
                self.data_type = 'long'
            else:
                self.data_type = 'wide'
                self.last_filtered_data = df.copy()
        else:
            self.data_type = 'long'

        logger.debug(f"Data type detected: {self.data_type}")
        self.show_processed_data()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)

        controls_group = QGroupBox("Table Controls")
        controls_layout = QHBoxLayout(controls_group)
        controls_layout.setSpacing(12)
        controls_layout.setContentsMargins(10, 10, 10, 10)

        search_label = QLabel("Search:")
        search_label.setFont(QFont("Segoe UI", 12))
        controls_layout.addWidget(search_label)
        self.search_entry = QLineEdit()
        self.search_entry.setPlaceholderText("Enter search term...")
        self.search_entry.setToolTip("Enter text to search in the pivot table")
        self.search_entry.setFixedWidth(200)
        self.search_entry.textChanged.connect(self.debounce_search)
        controls_layout.addWidget(self.search_entry)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedWidth(200)
        self.progress_bar.setVisible(False)
        controls_layout.addWidget(self.progress_bar)

        self.search_timer = QTimer(self)
        self.search_timer.setSingleShot(True)
        self.search_timer.timeout.connect(self.show_processed_data)

        filter_button = QPushButton("Filter")
        filter_button.setToolTip("Filter the pivot table by Solution Label or Element")
        filter_button.setMinimumWidth(120)
        filter_button.clicked.connect(self.open_filter_window)
        controls_layout.addWidget(filter_button)

        clear_col_filters_btn = QPushButton("Clear Col Filters")
        clear_col_filters_btn.setToolTip("Clear all column filters")
        clear_col_filters_btn.setMinimumWidth(120)
        clear_col_filters_btn.clicked.connect(self.clear_column_filters)
        controls_layout.addWidget(clear_col_filters_btn)

        self.save_button = QPushButton("Save Excel")
        self.save_button.setToolTip("Save the pivot table to an Excel file")
        self.save_button.clicked.connect(self.save_processed_excel)
        self.save_button.setMinimumWidth(120)
        controls_layout.addWidget(self.save_button)

        self.save_raw_button = QPushButton("Save Raw Excel")
        self.save_raw_button.setToolTip("Save the raw table without pivoting")
        self.save_raw_button.clicked.connect(self.save_raw_excel)
        self.save_raw_button.setMinimumWidth(150)
        controls_layout.addWidget(self.save_raw_button)

        self.decimal_combo = QComboBox()
        self.decimal_combo.addItems(["0", "1", "2", "3"])
        self.decimal_combo.setCurrentText(self.decimal_places)
        self.decimal_combo.setFixedWidth(60)
        self.decimal_combo.setToolTip("Set the number of decimal places for numeric values")
        self.decimal_combo.currentTextChanged.connect(self.show_processed_data)

        report_button = QPushButton("Report Changes")
        report_button.clicked.connect(self.show_changes_report)
        controls_layout.addWidget(report_button)

        compare_button = QPushButton("Compare Rows")
        compare_button.setToolTip("Compare two selected rows and compute differences")
        compare_button.clicked.connect(self.compare_selected_rows)
        controls_layout.addWidget(compare_button)

        similar_button = QPushButton("Find Similar")
        similar_button.setToolTip("Find rows similar to the selected row")
        similar_button.clicked.connect(self.find_similar_rows)
        controls_layout.addWidget(similar_button)

        oreas_button = QPushButton("Compare with OREAS")
        oreas_button.setToolTip("Compare selected rows with OREAS in Find Similarity tab")
        oreas_button.clicked.connect(self.compare_with_oreas)
        controls_layout.addWidget(oreas_button)

        controls_layout.addWidget(self.decimal_combo)
        controls_layout.addStretch()

        layout.addWidget(controls_group)

        table_group = QGroupBox("Pivot Table")
        table_layout = QVBoxLayout(table_group)
        table_layout.setContentsMargins(0, 0, 0, 0)

        self.processed_table = FreezeTableWidget(PandasModel(), frozen_columns=2, parent=self)
        self.processed_table.set_header_click_callback(self.on_header_clicked)
        self.processed_table.setStyleSheet(global_style)
        self.processed_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.processed_table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self.processed_table.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.processed_table.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOn)
        self.processed_table.setToolTip("Processed pivot table with filtered data")
        self.processed_table.setEnabled(False)
        self.processed_table.horizontalHeader().sectionClicked.connect(lambda section: self.on_header_clicked(section))
        self.processed_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        table_layout.addWidget(self.processed_table)

        layout.addWidget(table_group, stretch=1)
        self.setLayout(layout)

    def get_selected_checkbox_rows(self):
        model = self.processed_table.model()
        if model is None:
            return []
        return [i for i in range(model.rowCount()) if model._checkboxes[i]]

    def compare_with_oreas(self):
        selected_rows = self.get_selected_checkbox_rows()
        if len(selected_rows) < 1:
            QMessageBox.warning(self, "Warning", "Please select at least one row using checkboxes to compare with OREAS.")
            return

        df = self.last_filtered_data
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No data available for comparison.")
            return

        control_data = {'SAMPLE ID': []}
        for row_idx in selected_rows:
            selected_row = df.iloc[row_idx]
            solution_label = selected_row['Solution Label']
            control_data['SAMPLE ID'].append(solution_label)
            for col in df.columns:
                if col != 'Solution Label':
                    if col not in control_data:
                        control_data[col] = []
                    control_data[col].append(selected_row[col])

        control_df = pd.DataFrame(control_data)
        self.app.compare_tab.set_control_from_results(control_df)

        if hasattr(self.app, 'main_content'):
            self.app.main_content.switch_tab("Find similarity")
        else:
            QMessageBox.warning(self, "Error", "Cannot switch to Find Similarity tab.")

        QMessageBox.information(self, "Success", f"Selected {len(selected_rows)} row(s) sent to Find Similarity tab as control.")

    def show_changes_report(self):
        dialog = ChangesReportDialog(self.app, self)
        dialog.exec()

    def connect_to_crm_check(self, crm_check):
        logger.debug(f"Connecting CrmCheck to ResultsFrame instance_id: {self.instance_id}")
        crm_check.data_changed.connect(self.on_crm_data_changed)

    def on_crm_data_changed(self):
        logger.debug(f"Data changed in CrmCheck, updating ResultsFrame table instance_id: {self.instance_id}")
        self.update_table(self.last_filtered_data)

    def debounce_search(self, text):
        self.search_var = text
        self.search_timer.start(500)

    def format_value(self, x):
        try:
            value = float(x)
            decimal_places = int(self.decimal_combo.currentText())
            formatted = f"{value:.{decimal_places}f}".rstrip('0').rstrip('.')
            return formatted
        except (ValueError, TypeError):
            return str(x)

    def is_numeric(self, value):
        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False

    def reset_filter_cache(self):
        self.last_filtered_data = None
        self._last_cache_key = None
        self.data_hash = None
        self.column_filters.clear()
        self.filter_values.clear()
        logger.debug(f"Reset filter cache for instance_id: {self.instance_id}")

    def data_changed(self):
        logger.debug(f"ResultsFrame data_changed for instance_id: {self.instance_id}")
        self.reset_filter_cache()
        self.last_pivot_data = None
        self.show_processed_data()

    def on_data_changed(self):
        logger.debug(f"on_data_changed triggered for instance_id: {self.instance_id}")
        self.reset_filter_cache()
        self.last_pivot_data = None
        self.show_processed_data()

    def get_filter_cache_key(self):
        search = self.search_var.lower().strip()
        filters = str(self.column_filters)
        values = str(self.filter_values)
        return (search, filters, values)

    def apply_filters_to_wide_data(self, df):
        """Apply search and column filters to wide-format (pivoted) data"""
        filtered = df.copy()

        # Search filter
        search_text = self.search_var.lower().strip()
        if search_text:
            mask = filtered['Solution Label'].astype(str).str.lower().str.contains(search_text, na=False, regex=False)
            filtered = filtered[mask]

        # Column filters
        for col_name, col_filter in self.column_filters.items():
            if col_name not in filtered.columns:
                continue
            col_data = pd.to_numeric(filtered[col_name], errors='coerce')
            is_numeric_col = col_data.notna().any()

            if 'min_val' in col_filter and col_filter['min_val'] is not None and is_numeric_col:
                filtered = filtered[(col_data >= col_filter['min_val']) | col_data.isna()]
            if 'max_val' in col_filter and col_filter['max_val'] is not None and is_numeric_col:
                filtered = filtered[(col_data <= col_filter['max_val']) | col_data.isna()]
            if 'selected_values' in col_filter and col_filter['selected_values']:
                if is_numeric_col:
                    vals = {float(v) for v in col_filter['selected_values'] if self.is_numeric(v)}
                    filtered = filtered[col_data.isin(vals)]
                else:
                    filtered = filtered[filtered[col_name].isin(col_filter['selected_values'])]

        # Solution Label filter
        if self.filter_field == 'Solution Label':
            selected = [k for k, v in self.filter_values.get('Solution Label', {}).items() if v]
            if selected:
                filtered = filtered[filtered['Solution Label'].isin(selected)]

        return filtered.reset_index(drop=True)

    def compute_filtered_data(self):
        logger.debug(f"Starting compute_filtered_data (ResultsFrame) - data_type: {getattr(self, 'data_type', 'long')}")

        df = self.app.get_data()
        if df is None or len(df) == 0:
            return pd.DataFrame()

        # اگر داده قبلاً پیوت شده (wide) باشد → فقط فیلتر اعمال کن
        if getattr(self, 'data_type', 'long') == 'wide':
            logger.debug("Wide format detected – skipping pivot, only applying filters")
            return self.apply_filters_to_wide_data(df)

        # اگر long format باشد → باید پیوت کنیم (با NumPy خالص!)
        required = ['Solution Label', 'Element', 'Corr Con', 'Type']
        if not all(col in df.columns for col in required):
            logger.error("Missing required columns for pivot")
            return pd.DataFrame()

        try:
            # فقط نمونه‌ها
            samples = df[df['Type'].isin(['Samp', 'Sample'])].copy()
            if samples.empty:
                return pd.DataFrame()

            # exclusions
            exclude_samples = self.app.get_excluded_samples()
            exclude_volumes = self.app.get_excluded_volumes()
            exclude_dfs = self.app.get_excluded_dfs()
            samples = samples[
                (~samples['Solution Label'].isin(exclude_samples)) &
                (~samples['Solution Label'].isin(exclude_volumes)) &
                (~samples['Solution Label'].isin(exclude_dfs))
            ]
            if samples.empty:
                return pd.DataFrame()

            # حفظ ترتیب اصلی
            samples = samples.reset_index(drop=True)
            samples['original_index'] = samples.index

            # تمیزکاری
            samples['Solution Label'] = samples['Solution Label'].fillna('Unknown').astype(str)
            samples['Solution Label'] = samples['Solution Label'].str.replace(r'(?i)^nan$', 'Unknown', regex=True)
            samples['Element'] = samples['Element'].astype(str).str.split('_').str[0]
            samples['Corr Con'] = pd.to_numeric(samples['Corr Con'], errors='coerce')

            records = samples.to_dict('records')

            # گروه‌بندی بر اساس Solution Label
            solution_groups = {}
            for r in records:
                sl = r['Solution Label']
                solution_groups.setdefault(sl, []).append(r)

            # محاسبه set_size
            most_common_sizes = {}
            for sl, group in solution_groups.items():
                counts = {}
                for r in group:
                    counts[r['Element']] = counts.get(r['Element'], 0) + 1
                values = list(counts.values())
                g = reduce(math.gcd, values) if values else 1
                total = len(group)
                most_common_sizes[sl] = total // g if g > 1 and total % g == 0 else total

            # تشخیص تکرار
            has_repeats = False
            check = {}
            for r in records:
                sl = r['Solution Label']
                pos = next(i for i, x in enumerate(solution_groups[sl]) if x is r)
                gid_approx = pos // most_common_sizes[sl]
                key = (sl, gid_approx, r['Element'])
                check[key] = check.get(key, 0) + 1
                if check[key] > 1:
                    has_repeats = True
                    break

            # تابع مرتب‌سازی برچسب
            def label_key(x):
                s = str(x).replace(' ', '')
                m = re.search(r'(\d+)', s)
                return (s.lower() if not m else s[:m.start()].lower(), int(m.group(1)) if m else 0)

            final_rows = []

            if has_repeats:
                # حالت تکرار → گروه‌بندی دقیق
                for r in records:
                    sl = r['Solution Label']
                    pos = next(i for i, x in enumerate(solution_groups[sl]) if x is r)
                    r['_group_id'] = pos // most_common_sizes[sl]

                occ_count = {}
                for r in records:
                    k = (r['Solution Label'], r['_group_id'], r['Element'])
                    occ_count[k] = occ_count.get(k, 0) + 1

                occ_counter = {}
                for r in records:
                    k = (r['Solution Label'], r['_group_id'], r['Element'])
                    n = occ_count[k]
                    idx = occ_counter.get(k, 0) + 1
                    occ_counter[k] = idx
                    r['_col'] = f"{r['Element']}_{idx}" if n > 1 else r['Element']

                row_map = {}
                for r in records:
                    rid = (r['Solution Label'], r['_group_id'])
                    row_map.setdefault(rid, {'Solution Label': r['Solution Label']})
                    row_map[rid][r['_col']] = r['Corr Con']

                # مرتب‌سازی بر اساس اولین ایندکس
                def first_idx(rid):
                    sl, gid = rid
                    return min(x['original_index'] for x in records if x['Solution Label'] == sl and x.get('_group_id') == gid)

                ordered = sorted(row_map.items(), key=lambda x: first_idx(x[0]))
                final_rows = [row for _, row in ordered]

                # تعیین ترتیب عناصر از اولین گروه کامل
                first_full = next((row for _, row in row_map.items()
                                if len(row) - 1 >= most_common_sizes.get(row['Solution Label'], 1)), None)
                element_order = sorted(
                    [k for k in (first_full or final_rows[0]).keys() if k != 'Solution Label'],
                    key=label_key
                )
                self.element_order = element_order

            else:
                # بدون تکرار
                uid_map = {}
                for r in records:
                    k = (r['Solution Label'], r['Element'])
                    uid_map[k] = uid_map.get(k, -1) + 1
                    r['_uid'] = uid_map[k]

                row_map = {}
                for r in records:
                    rid = (r['Solution Label'], r['_uid'])
                    row_map.setdefault(rid, {'Solution Label': r['Solution Label']})
                    row_map[rid][r['Element']] = r['Corr Con']

                def first_idx(rid):
                    sl, uid = rid
                    return min((x['original_index'] for x in records if x['Solution Label'] == sl and x.get('_uid') == uid), default=999999)

                ordered = sorted(row_map.items(), key=lambda x: first_idx(x[0]))
                final_rows = [row for _, row in ordered]

                cols = set()
                for r in final_rows:
                    cols.update(k for k in r if k != 'Solution Label')
                self.element_order = sorted(cols, key=label_key)

            # مرتب‌سازی Solution Label
            self.solution_label_order = sorted(
                {r['Solution Label'] for r in final_rows},
                key=label_key
            )

            # ساخت DataFrame نهایی
            if not final_rows:
                return pd.DataFrame(columns=['Solution Label'])

            pivot_df = pd.DataFrame(final_rows)

            # اعمال ترتیب ستون‌ها
            if self.element_order:
                cols = ['Solution Label'] + [c for c in self.element_order if c in pivot_df.columns]
                missing = [c for c in pivot_df.columns if c not in cols]
                pivot_df = pivot_df[cols + missing]

            # کش کردن
            self.last_pivot_data = pivot_df
            self.data_hash = str(pd.util.hash_pandas_object(samples).sum())

            # حالا فیلترها رو اعمال کن
            filtered = self.apply_filters_to_wide_data(pivot_df)
            self.last_filtered_data = filtered
            return filtered

        except Exception as e:
            logger.error(f"Pivot error in ResultsFrame: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    @pyqtSlot(dict)
    def update_results_from_compare(self, updates):
        if not hasattr(self, 'processed_table') or not self.processed_table:
            return
        model = self.processed_table.model()
        if not model or not isinstance(model, PandasModel):
            return

        updated = 0
        for control_id, col_updates in updates.items():
            sample_id_col = 1
            found_row = -1
            for row in range(model.rowCount()):
                index = model.index(row, sample_id_col)
                current_id = model.data(index, Qt.ItemDataRole.DisplayRole)
                if current_id == control_id:
                    found_row = row
                    break
            if found_row == -1:
                continue
            for col_name, new_val in col_updates.items():
                col_idx = -1
                for c in range(model.columnCount()):
                    header_text = model.headerData(c, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                    if header_text == col_name:
                        col_idx = c
                        break
                if col_idx == -1:
                    continue
                index = model.index(found_row, col_idx)
                success = model.setData(index, f"{new_val:.2f}", Qt.ItemDataRole.EditRole)
                if success:
                    updated += 1

        if updated > 0:
            self.processed_table.viewport().update()

    def show_processed_data(self):
        if self.worker is not None and self.worker.isRunning():
            return

        # For wide format, update directly
        if getattr(self, 'data_type', 'long') == 'wide':
            if self.last_filtered_data is not None:
                self.update_table(self.last_filtered_data)
            return

        # For long format, use worker
        self.progress_bar.setVisible(True)
        self.progress_bar.setRange(0, 0)
        self.search_entry.setEnabled(False)
        self.worker = DataWorker(self)
        self.worker.data_ready.connect(self.update_table)
        self.worker.error_occurred.connect(self.show_error)
        self.worker.finished.connect(self.on_worker_finished)
        self.worker.start()

    def on_worker_finished(self):
        self.progress_bar.setVisible(False)
        self.search_entry.setEnabled(True)
        if hasattr(self.app, 'notify_data_changed'):
            self.app.notify_data_changed()

    def update_table(self, df):
        logger.debug(f"Updating table for instance_id: {self.instance_id}, data shape: {df.shape if df is not None else 'None'}")
        if df is None or df.empty:
            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["Status"])
            model.appendRow([QStandardItem("No data available after filtering")])
            self.processed_table.setModel(model)
            self.processed_table.setColumnWidth(0, 150)
            self.column_widths = {"Status": 150}
            self.processed_table.frozenTableView.setModel(model)
            self.processed_table.update_frozen_columns()
            self.processed_table.setEnabled(False)
            return

        columns = list(df.columns)
        self.column_widths = {}
        self.processed_table.setColumnWidth(0, 50)
        self.column_widths["Select"] = 50
        self.processed_table.setColumnWidth(1, 150)
        self.column_widths["Solution Label"] = 150

        for col_idx, col in enumerate(columns, 1):
            if col == 'Solution Label':
                continue
            sample_data = df[col].dropna().astype(str)
            max_width = max([len(str(col))] + [len(str(self.format_value(x))) for x in sample_data], default=10)
            pixel_width = min(max_width * 10, 300)
            self.column_widths[col] = pixel_width
            self.processed_table.setColumnWidth(col_idx, pixel_width)


        df['file_name'] = None
        for file in self.app.file_ranges:
            start = file['start_pivot_row']
            end = file['end_pivot_row']
            clean_name = file['clean_name']
            mask = df.index[(df.index >= start) & (df.index <= end)]
            if len(mask) > 0:
                df.loc[mask, 'file_name'] = clean_name
                
        model = PandasModel(df, format_value=self.format_value)
        self.processed_table.setModel(model)
        self.processed_table.frozenTableView.setModel(model)
        self.processed_table.update_frozen_columns()
        
        self.processed_table.model().layoutChanged.emit()
        self.processed_table.frozenTableView.model().layoutChanged.emit()
        self.processed_table.viewport().update()
        self.processed_table.frozenTableView.viewport().update()
        self.processed_table.setEnabled(True)
        
        if hasattr(self.app, 'notify_data_changed'):
            self.app.notify_data_changed()

    def show_error(self, message):
        self.progress_bar.setVisible(False)
        self.search_entry.setEnabled(True)
        QMessageBox.warning(self, "Error", message)

    def open_filter_window(self):
        df = self.app.get_data()
        if df is None:
            QMessageBox.warning(self, "Warning", "No data to filter!")
            return

    def on_header_clicked(self, section, col_name=None):
        if col_name is None:
            model = self.processed_table.model()
            if model is not None:
                col_name = model.headerData(section, Qt.Orientation.Horizontal, Qt.ItemDataRole.DisplayRole)
                col_name = str(col_name) if col_name is not None else f"Column {section}"
            else:
                return

        if col_name == "Select":
            return

        data_source = self.last_filtered_data
        if data_source is None or data_source.empty:
            QMessageBox.warning(self, "Warning", "No data available to filter!")
            return

        if col_name not in data_source.columns:
            return

        dialog = ColumnFilterDialog(
            parent=self,
            col_name=col_name,
            data_source=data_source,
            column_filters=self.column_filters,
            on_apply_callback=self.show_processed_data
        )
        dialog.exec()

    def clear_column_filters(self):
        self.column_filters.clear()
        self.show_processed_data()
        QMessageBox.information(self, "Filters Cleared", "All column filters have been cleared.")

    def compare_selected_rows(self):
        selected_rows = self.get_selected_checkbox_rows()
        if len(selected_rows) != 2:
            QMessageBox.warning(self, "Warning", "Please select exactly two rows using checkboxes to compare.")
            return

        df = self.last_filtered_data
        if df is None or df.empty:
            return

        row1 = df.iloc[selected_rows[0]]
        row2 = df.iloc[selected_rows[1]]
        label1 = row1['Solution Label']
        label2 = row2['Solution Label']

        diff_df = pd.DataFrame({
            'Column': df.columns,
            'Row1 Value': row1.values,
            'Row2 Value': row2.values,
            'Difference': np.nan
        })

        for col in df.columns:
            if col != 'Solution Label':
                df[col] = pd.to_numeric(df[col], errors='coerce')

        numeric_cols = [col for col in df.columns if pd.api.types.is_numeric_dtype(df[col]) and col != 'Solution Label']

        for col in numeric_cols:
            try:
                val1 = float(row1[col]) if pd.notna(row1[col]) else np.nan
                val2 = float(row2[col]) if pd.notna(row2[col]) else np.nan
                if pd.notna(val1) and pd.notna(val2):
                    diff = val1 - val2
                    diff_df.loc[diff_df['Column'] == col, 'Difference'] = diff
            except (ValueError, TypeError):
                pass

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Comparison: {label1} vs {label2}")
        dialog_layout = QVBoxLayout(dialog)
        diff_table = QTableView()
        diff_model = PandasModel(diff_df, format_value=self.format_value)
        diff_table.setModel(diff_model)
        diff_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        dialog_layout.addWidget(diff_table)
        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.close)
        dialog_layout.addWidget(close_button)
        dialog.resize(800, 600)
        dialog.exec()

    def find_similar_rows(self):
        selected_rows = self.get_selected_checkbox_rows()
        if len(selected_rows) < 1:
            QMessageBox.warning(self, "Warning", "Please select at least one row using checkboxes to find similar rows.")
            return

        row_idx = selected_rows[0]
        df = self.last_filtered_data
        if df is None or df.empty:
            return

        selected_row = df.iloc[row_idx]
        label = selected_row['Solution Label']

        data_numeric = df.copy()
        for col in data_numeric.columns:
            if col != 'Solution Label':
                data_numeric[col] = pd.to_numeric(data_numeric[col], errors='coerce')

        numeric_cols = [col for col in data_numeric.columns if pd.api.types.is_numeric_dtype(data_numeric[col]) and col != 'Solution Label']

        if not numeric_cols:
            QMessageBox.warning(self, "Warning", "No numeric columns available for similarity computation.")
            return

        data_numeric = data_numeric[numeric_cols].fillna(0).astype(float)
        selected_vec = data_numeric.iloc[row_idx].values
        distances = np.linalg.norm(data_numeric.values - selected_vec, axis=1)
        df['Distance'] = distances
        similar_df = df.sort_values('Distance').reset_index(drop=True)
        similar_df = similar_df.drop(columns=['Distance'])

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Similar Rows to {label}")
        dialog_layout = QVBoxLayout(dialog)
        similar_table = QTableView()
        similar_model = PandasModel(similar_df, format_value=self.format_value)
        similar_table.setModel(similar_model)
        similar_table.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        dialog_layout.addWidget(similar_table)
        close_button = QPushButton("Close")
        close_button.clicked.connect(dialog.close)
        dialog_layout.addWidget(close_button)
        dialog.resize(800, 600)
        dialog.exec()

    def save_processed_excel(self):
        df = self.last_filtered_data
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No data to save!")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if file_path:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Processed Pivot Table"

                header_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                first_column_fill = PatternFill(start_color="FFF5E4", end_color="FFF5E4", fill_type="solid")
                odd_row_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                even_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                header_font = Font(name="Segoe UI", size=12, bold=True)
                cell_font = Font(name="Segoe UI", size=12)
                cell_alignment = Alignment(horizontal="center", vertical="center")
                thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

                headers = list(df.columns)
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = cell_alignment
                    cell.border = thin_border
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15

                for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                    fill = even_row_fill if (row_idx - 1) % 2 == 0 else odd_row_fill
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if pd.isna(value):
                            cell.value = None
                        else:
                            try:
                                float_value = float(value)
                                cell.value = float_value
                                decimal_places = int(self.decimal_combo.currentText())
                                # cell.number_format = f"0.{'0' * decimal_places}"
                                # cell.value = self.format_value(value)
                            except ValueError:
                                cell.value = str(value)
                        cell.font = cell_font
                        cell.alignment = cell_alignment
                        cell.border = thin_border
                        if col_idx == 1:
                            cell.fill = first_column_fill
                        else:
                            cell.fill = fill

                wb.save(file_path)
                QMessageBox.information(self, "Success", "Processed pivot table saved successfully!")

                if QMessageBox.question(self, "Open File", "Would you like to open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                    try:
                        system = platform.system()
                        if system == "Windows":
                            os.startfile(file_path)
                        elif system == "Darwin":
                            os.system(f"open {file_path}")
                        else:
                            os.system(f"xdg-open {file_path}")
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save: {str(e)}")

    def save_raw_excel(self):
        df = self.app.get_data()
        if df is None or df.empty:
            QMessageBox.warning(self, "Warning", "No raw data to save!")
            return

        df_filtered = df[df['Type'].isin(['Samp', 'Sample'])].copy()
        df_filtered = df_filtered[
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_samples())) &
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_volumes())) &
            (~df_filtered['Solution Label'].isin(self.app.get_excluded_dfs()))
        ]

        if df_filtered.empty:
            QMessageBox.warning(self, "Warning", "No filtered raw data to save!")
            return

        file_path, _ = QFileDialog.getSaveFileName(self, "Save Raw Excel File", "", "Excel Files (*.xlsx)")
        if file_path:
            try:
                df_filtered.to_excel(file_path, index=False)
                QMessageBox.information(self, "Success", "Raw table saved successfully!")

                if QMessageBox.question(self, "Open File", "Would you like to open the saved Excel file?") == QMessageBox.StandardButton.Yes:
                    try:
                        system = platform.system()
                        if system == "Windows":
                            os.startfile(file_path)
                        elif system == "Darwin":
                            os.system(f"open {file_path}")
                        else:
                            os.system(f"xdg-open {file_path}")
                    except Exception as e:
                        QMessageBox.critical(self, "Error", f"Failed to open file: {str(e)}")

            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save raw data: {str(e)}")

    def reset_cache(self):
        self.last_filtered_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.column_widths = {}
        self.filter_values = {}
        self.column_filters = {}
        self.search_var = ""
        self.data_hash = None

    def reset_state(self):
        self.search_var = ""
        self.filter_field = "Solution Label"
        self.filter_values = {}
        self.column_filters = {}
        self.column_widths = {}
        self.column_backups = {}
        self.last_filtered_data = None
        self.last_pivot_data = None
        self._last_cache_key = None
        self.solution_label_order = None
        self.element_order = None
        self.decimal_places = "1"
        self.data_hash = None

        if hasattr(self, 'search_entry'):
            self.search_entry.setText("")
        if hasattr(self, 'decimal_combo'):
            self.decimal_combo.setCurrentText(self.decimal_places)
        if hasattr(self, 'processed_table'):
            model = QStandardItemModel()
            model.setHorizontalHeaderLabels(["Status"])
            model.appendRow([QStandardItem("No data available after filtering")])
            self.processed_table.setModel(model)
            self.processed_table.setColumnWidth(0, 150)
            self.processed_table.frozenTableView.setModel(model)
            self.processed_table.update_frozen_columns()
            self.processed_table.setEnabled(False)

        if self.worker is not None and self.worker.isRunning():
            self.worker.terminate()
            self.worker = None
            self.progress_bar.setVisible(False)
            self.search_entry.setEnabled(True)

        if hasattr(self.app, 'notify_data_changed'):
            self.app.notify_data_changed()