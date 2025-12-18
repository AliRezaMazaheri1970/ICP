from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableView, QHeaderView, 
    QGroupBox, QMessageBox, QLineEdit, QLabel, QComboBox, QFileDialog,
    QDialog, QRadioButton, QCheckBox, QButtonGroup, QDialogButtonBox
)
from PyQt6.QtCore import pyqtSignal
import pandas as pd
import re
import logging
import sqlite3
from ...Common.Freeze_column import FreezeTableWidget
from ..pivot_table_model import PivotTableModel
from ..verification.crm_calibration import PivotPlotWindow
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from styles.common import common_styles
from screens.process.crm.crm_manager import CRMManager
# Setup logging
logging.basicConfig(level=logging.DEBUG, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Global stylesheet
class CrmCheck(QWidget):
    data_changed = pyqtSignal()
    
    def __init__(self, app, results_frame, parent=None):
        super().__init__(parent)
        self.app = app
        self.results_frame = results_frame
        self.corrected_crm = {}
        self._inline_crm_rows = {}  # Now uses unique key: {(label, row_index): [crm_data]}
        self._inline_crm_rows_display = {}
        self.included_crms = {}
        self.column_widths = {}
        self.crm_manager = CRMManager(self)
        self.crm_diff_min = QLineEdit("-12")
        self.crm_diff_max = QLineEdit("12")
        self.current_plot_window = None
        self.setup_ui()
        self.results_frame.app.notify_data_changed = self.on_data_changed
        if hasattr(self.results_frame, 'decimal_combo') and self.results_frame.decimal_combo is not None:
            self.results_frame.decimal_combo.currentTextChanged.connect(self.update_pivot_display)
        else:
            logger.error("decimal_combo is not defined in ResultsFrame")
            raise AttributeError("decimal_combo is not defined in ResultsFrame")

    def setup_ui(self):
        """Set up the UI with styling matching EmptyCheckFrame."""
        self.setStyleSheet(common_styles)
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(15)

        # Control group
        control_group = QGroupBox("Pivot Controls")
        control_layout = QHBoxLayout(control_group)
        control_layout.setSpacing(10)

        # Check CRM button
        check_crm_btn = QPushButton("Check CRM")
        check_crm_btn.setMinimumWidth(80)
        check_crm_btn.clicked.connect(self.crm_manager.check_rm)
        control_layout.addWidget(check_crm_btn)

        # Manual CRM button
        manual_crm_btn = QPushButton("Manual CRM")
        manual_crm_btn.setMinimumWidth(80)
        manual_crm_btn.clicked.connect(self.manual_crm_selection)
        control_layout.addWidget(manual_crm_btn)

        # CRM Range input
        crm_range_label = QLabel("CRM Range (%):")
        control_layout.addWidget(crm_range_label)
        self.crm_diff_min.setFixedWidth(50)
        self.crm_diff_min.textChanged.connect(self.validate_crm_diff_range)
        control_layout.addWidget(self.crm_diff_min)
        control_layout.addWidget(QLabel("to"))
        self.crm_diff_max.setFixedWidth(50)
        self.crm_diff_max.textChanged.connect(self.validate_crm_diff_range)
        control_layout.addWidget(self.crm_diff_max)

        # Decimal places input
        decimal_label = QLabel("Decimal Places:")
        control_layout.addWidget(decimal_label)
        if hasattr(self.results_frame, 'decimal_combo') and self.results_frame.decimal_combo is not None:
            self.results_frame.decimal_combo.setFixedWidth(60)
            self.results_frame.decimal_combo.setToolTip("Set the number of decimal places for numeric values")
            control_layout.addWidget(self.results_frame.decimal_combo)
        else:
            logger.warning("decimal_combo not available, creating a local one")
            local_decimal_combo = QComboBox()
            local_decimal_combo.addItems(["0", "1", "2", "3"])
            local_decimal_combo.setCurrentText("1")
            local_decimal_combo.setFixedWidth(60)
            local_decimal_combo.setToolTip("Set the number of decimal places for numeric values")
            local_decimal_combo.currentTextChanged.connect(self.update_pivot_display)
            control_layout.addWidget(local_decimal_combo)

        # Clear CRM button
        clear_crm_btn = QPushButton("Clear CRM")
        clear_crm_btn.setMinimumWidth(80)
        clear_crm_btn.clicked.connect(self.clear_inline_crm)
        control_layout.addWidget(clear_crm_btn)

        # Calib button
        calib_btn = QPushButton("Calib")
        calib_btn.setMinimumWidth(80)
        calib_btn.clicked.connect(self.show_element_plot)
        control_layout.addWidget(calib_btn)

        # Export button
        export_btn = QPushButton("Export")
        export_btn.setMinimumWidth(80)
        export_btn.clicked.connect(self.export_table)
        control_layout.addWidget(export_btn)

        control_layout.addStretch()
        main_layout.addWidget(control_group)

        # Table group
        table_group = QGroupBox("Pivot Table")
        table_layout = QVBoxLayout(table_group)
        table_layout.setSpacing(10)

        # Table view
        self.table_view = FreezeTableWidget(PivotTableModel(self))
        self.table_view.setAlternatingRowColors(True)
        self.table_view.setSelectionBehavior(QTableView.SelectionBehavior.SelectRows)
        self.table_view.setSortingEnabled(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        table_layout.addWidget(self.table_view)

        main_layout.addWidget(table_group, stretch=1)

    def manual_crm_selection(self):
        """Open a dialog to manually select a CRM for a selected row."""
        selected_rows = self.table_view.selectionModel().selectedRows()
        if not selected_rows:
            QMessageBox.warning(self, "Warning", "Please select a row to assign a CRM!")
            logger.warning("No row selected for manual CRM assignment")
            return

        if len(selected_rows) > 1:
            QMessageBox.warning(self, "Warning", "Please select only one row!")
            logger.warning("Multiple rows selected for manual CRM assignment")
            return

        row_index = selected_rows[0].row()
        model = self.table_view.model()
        if not model:
            QMessageBox.warning(self, "Warning", "No data available in table!")
            logger.warning("No data available in table for manual CRM")
            return

        solution_label = model.data(model.index(row_index, 0))
        if not solution_label:
            QMessageBox.warning(self, "Warning", "Invalid row selected!")
            logger.warning("Invalid row selected for manual CRM")
            return

        self.crm_manager.open_manual_crm_dialog(solution_label, row_index)
        self.update_pivot_display()
        self.data_changed.emit()

    def on_data_changed(self):
        """Update pivot table when data in ResultsFrame changes."""
        logger.debug("Data changed in ResultsFrame, updating pivot display")
        self.update_pivot_display()

    def validate_crm_diff_range(self):
        """Validate CRM difference range inputs and update display."""
        try:
            min_val = float(self.crm_diff_min.text())
            max_val = float(self.crm_diff_max.text())
            if min_val > max_val:
                self.crm_diff_min.setText(str(max_val - 1))
            logger.debug(f"CRM diff range set to {min_val} to {max_val}")
            self.crm_manager.check_rm_with_diff_range(min_val, max_val)
            self.update_pivot_display()
            self.data_changed.emit()
        except ValueError:
            logger.debug("Invalid CRM diff range input, skipping update")
            pass

    def update_pivot_display(self):
        """Update the pivot table display using ResultsFrame's last_filtered_data."""
        logger.debug("Starting update_pivot_display")
        pivot_data = self.results_frame.last_filtered_data

        if pivot_data is None or pivot_data.empty:
            logger.warning("No data loaded for pivot display")
            self.table_view.setModel(None)
            self.table_view.frozenTableView.setModel(None)
            return

        logger.debug(f"Current view data shape: {pivot_data.shape}")
        self._inline_crm_rows_display = self.crm_manager._build_crm_row_lists_for_columns(list(pivot_data.columns))
        
        # Build combined_rows with proper tuple keys
        combined_rows = []
        for idx, sol_label in enumerate(pivot_data['Solution Label']):
            row_key = (sol_label, idx)
            if row_key in self._inline_crm_rows_display:
                combined_rows.append((row_key, self._inline_crm_rows_display[row_key]))

        model = PivotTableModel(self, pivot_data, combined_rows)
        self.table_view.setModel(model)
        self.table_view.frozenTableView.setModel(model)
        self.table_view.update_frozen_columns()
        self.table_view.model().layoutChanged.emit()
        self.table_view.frozenTableView.model().layoutChanged.emit()
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        for col, width in self.column_widths.items():
            if col < len(pivot_data.columns):
                self.table_view.horizontalHeader().resizeSection(col, width)
        self.table_view.viewport().update()
        logger.debug("Completed update_pivot_display")

    def clear_inline_crm(self):
        """Clear inline CRM data."""
        logger.debug("Clearing inline CRM data")
        self._inline_crm_rows.clear()
        self._inline_crm_rows_display.clear()
        self.included_crms.clear()
        self.update_pivot_display()
        self.data_changed.emit()

    def show_element_plot(self):
        """Show element calibration plot in a resizable window."""
        logger.debug("Attempting to show element plot")
        pivot_data = self.results_frame.last_filtered_data
        if pivot_data is None or pivot_data.empty:
            logger.warning("No data to plot")
            QMessageBox.warning(self, "Warning", "No data to plot!")
            return
        logger.debug("Opening element plot window")
        if hasattr(self, 'current_plot_window') and self.current_plot_window:
            self.current_plot_window.close()
        annotations = []
        self.current_plot_window = PivotPlotWindow(self, annotations)
        self.current_plot_window.show()

    def backup_column(self, column):
        """Backup a column before applying corrections."""
        pivot_data = self.results_frame.last_filtered_data
        if column in pivot_data.columns:
            self.results_frame.column_backups[column] = pivot_data[column].copy()
            logger.debug(f"Backed up column: {column}")
            self.data_changed.emit()

    def restore_column(self, column):
        """Restore a column from backup."""
        if column in self.results_frame.column_backups:
            self.results_frame.last_filtered_data[column] = self.results_frame.column_backups[column].copy()
            self.update_pivot_display()
            logger.debug(f"Restored column: {column}")
            del self.results_frame.column_backups[column]
            self.data_changed.emit()
        else:
            logger.warning(f"No backup found for column: {column}")
            QMessageBox.warning(self, "Warning", f"No backup found for column {column}")

    def export_table(self):
        """Export the pivot table with an optional colored CRM block."""
        logger.debug("Export started")
        pivot_data = self.results_frame.last_filtered_data
        if pivot_data is None or pivot_data.empty:
            QMessageBox.warning(self, "Warning", "No data to export!")
            return

        has_crm = bool(self._inline_crm_rows_display)

        export_with_style = False
        if has_crm:
            reply = QMessageBox.question(
                self,
                "Export CRM rows",
                "Colored CRM / Diff(%) rows are present.\n\n"
                "Do you want to export the full table with colors in Excel?\n"
                "• Yes – include CRM rows and apply colors\n"
                "• No  – export only the original pivot table (no extra rows, no colors)",
                QMessageBox.StandardButton.Yes |
                QMessageBox.StandardButton.No |
                QMessageBox.StandardButton.Cancel,
                QMessageBox.StandardButton.Yes
            )
            if reply == QMessageBox.StandardButton.Cancel:
                return
            export_with_style = (reply == QMessageBox.StandardButton.Yes)

        file_path, filter_ = QFileDialog.getSaveFileName(
            self,
            "Save Pivot Table",
            "",
            "Excel Files (*.xlsx);;CSV Files (*.csv);;All Files (*)"
        )
        if not file_path:
            return

        try:
            # ---------- CSV (no styling possible) ----------
            if filter_.startswith("CSV") or file_path.lower().endswith(".csv"):
                if has_crm and export_with_style:
                    QMessageBox.information(
                        self, "Info",
                        "CSV does not support colors or extra rows.\n"
                        "Only the original table will be exported."
                    )
                pivot_data.to_csv(file_path, index=True)
                QMessageBox.information(self, "Success", f"Exported to\n{file_path}")
                return

            # ---------- Excel ----------
            if not file_path.lower().endswith(".xlsx"):
                file_path += ".xlsx"

            # Build DataFrame
            if export_with_style and has_crm:
                df_to_export = self._build_full_table_with_crm_for_export()
            else:
                df_to_export = pivot_data.copy()

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                df_to_export.to_excel(writer, index=True, sheet_name="Pivot Table")

                if export_with_style and has_crm:
                    ws = writer.sheets["Pivot Table"]
                    self._apply_crm_styling_to_worksheet(ws, len(pivot_data))

            QMessageBox.information(self, "Success", f"Exported successfully to\n{file_path}")

        except Exception as e:
            logger.exception("Export failed")
            QMessageBox.critical(self, "Error", f"Export failed:\n{e}")


    def _build_full_table_with_crm_for_export(self):
        """Return a DataFrame that contains original rows + CRM rows + Diff rows."""
        pivot_data = self.results_frame.last_filtered_data
        rows = []

        try:
            dec = int(self.results_frame.decimal_combo.currentText())
        except Exception:
            dec = 1

        for idx in range(len(pivot_data)):
            orig_row = pivot_data.iloc[idx].to_dict()
            rows.append(orig_row)

            key = (pivot_data.iloc[idx]["Solution Label"], idx)
            blocks = self._inline_crm_rows_display.get(key, [])

            for block_idx, (values_list, tags) in enumerate(blocks):
                # CRM row
                crm_row = orig_row.copy()
                for col, val in zip(pivot_data.columns, values_list):
                    crm_row[col] = val if val != "" else None
                rows.append(crm_row)

                # Diff row (second block)
                if block_idx == 0 and len(blocks) > 1:
                    diff_values, diff_tags = blocks[1]
                    diff_row = orig_row.copy()
                    diff_row["Solution Label"] = f"{pivot_data.iloc[idx]['Solution Label']} Diff (%)"
                    for col, val in zip(pivot_data.columns, diff_values):
                        diff_row[col] = val if val not in ("", "N/A") else None
                    rows.append(diff_row)

        return pd.DataFrame(rows, columns=pivot_data.columns)


    def _apply_crm_styling_to_worksheet(self, ws, original_row_count):
        """Apply background colors to CRM and Diff rows in the Excel worksheet."""
        fill_crm        = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")  # light green
        fill_diff_base  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")  # light yellow
        fill_in_range   = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # green
        fill_out_range  = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # red

        excel_row = 2  # row 1 = header

        for idx in range(original_row_count):
            sol_label = self.results_frame.last_filtered_data.iloc[idx]["Solution Label"]
            key = (sol_label, idx)
            blocks = self._inline_crm_rows_display.get(key, [])

            # original data row – no fill
            excel_row += 1

            if not blocks:
                continue

            # ---- CRM row (first block) ----
            for cell in ws[excel_row]:
                cell.fill = fill_crm
            excel_row += 1

            # ---- Diff row (second block, if present) ----
            if len(blocks) > 1:
                _, diff_tags = blocks[1]
                for col_idx, tag in enumerate(diff_tags, start=1):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    if tag == "in_range":
                        cell.fill = fill_in_range
                    elif tag == "out_range":
                        cell.fill = fill_out_range
                    else:
                        cell.fill = fill_diff_base
                excel_row += 1

